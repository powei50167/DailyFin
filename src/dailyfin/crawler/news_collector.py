from datetime import datetime, timedelta
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests



def append_news(search_target, news_dict):
    now = datetime.now()

    # 週一抓兩天，其餘抓一天
    filterday = '2' if now.weekday() == 0 else '1'
    url = f'https://news.google.com/search?q={search_target}%20when%3A{filterday}d&hl=zh-TW&gl=TW&ceid=TW%3Azh-Hant'
    
    response = requests.get(url)
    if response.status_code != 200:
        print(f"{search_target}_HTTP請求失敗：{response.status_code}")
        return

    soup = BeautifulSoup(response.text, 'html.parser')
    articles = soup.select(".m5k28")

    # 時間過濾（週一 72 小時，其餘 24 小時）
    time_delta = timedelta(hours=72 if now.weekday() == 0 else 24)

    # 取得 news_dict 內的所有標題，避免重複加入
    existing_titles = {item['title'] for item in news_dict.values()}

    news_id = 0

    # 🔍 Google News 每篇文章主要容器
    for item in articles:

        title_tag = item.select_one("a.JtKRv")
        title = title_tag.text.strip()
        full_link = "https://news.google.com" + title_tag.get("href")[1:]

        # 若標題已存在則跳過
        if title in existing_titles:
            continue

        source_tag = item.select_one(".vr1PYe")

        if hasattr(source_tag, "text"):
            source = source_tag.text.strip()
        else:
            source = str(source_tag).strip() if source_tag else "No source"

        # -------------------------
        # 抓取時間（<time class="hvbAAd" datetime="...">）
        # -------------------------
        time_tag = item.find_next_sibling().select_one("time.hvbAAd")
        news_time = time_tag['datetime'] if time_tag and time_tag.has_attr('datetime') else None

        # 時間過濾
        if news_time:
            utc_time = datetime.strptime(news_time, '%Y-%m-%dT%H:%M:%SZ') + timedelta(hours=8)

            if now - utc_time <= time_delta:
                # -------------------------
                # 加入 news_dict
                # -------------------------
                news_dict[f'{search_target}_{news_id}'] = {
                    'title': title,
                    'link': full_link,
                    'news_time': utc_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'source': source
                }

                # 避免重複新增
                existing_titles.add(title)
                news_id += 1



def load_search_target(file_path: str) -> dict:
    news_dict = {}
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for search_target in row:
            if search_target:
                append_news(search_target.strip(), news_dict)
    print(f'新聞讀取完成,新聞總量 : {len(news_dict)}')
    return news_dict


load_search_target(r'C:\Users\rain50167\Desktop\PROJECT\DailyFin\search_targets.xlsx')